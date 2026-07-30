"""Tests for the hub-local graph and Excel export services.

These replaced ``metaseed.ui.services.graph``/``metaseed.ui.services.export``
(#53 step 4); they are built on the public ``MetaseedClient``/``ProfileFacade``
API and must produce the same shapes the templates and download route consume.
"""

from datetime import datetime

from metaseed import MetaseedClient
from openpyxl import load_workbook

from metaseed_hub.ui.services.export import export_to_bytes, generate_filename
from metaseed_hub.ui.services.graph import build_graph


def _facade_with_tree():
    """A miappe facade holding Investigation -> Study."""
    client = MetaseedClient("miappe", "1.1")
    inv = client.create_entity(
        "Investigation", {"unique_id": "inv-1", "title": "Trial"}, skip_validation=True
    )
    client.create_entity(
        "Study",
        {"unique_id": "st-1", "title": "Field study", "investigation_id": "inv-1"},
        parent_id=inv.id,
        skip_validation=True,
    )
    return client.facade


def test_build_graph_has_nodes_edges_and_full_entity_type_legend() -> None:
    facade = _facade_with_tree()

    graph = build_graph(facade)

    assert len(graph["nodes"]) == 2
    assert len(graph["edges"]) == 1
    # The legend lists every type the profile defines, not only instantiated ones.
    assert graph["entity_types"] == facade.entities
    assert "Person" in graph["entity_types"]


def test_export_creates_sheet_per_entity_type_with_rows() -> None:
    facade = _facade_with_tree()

    workbook = load_workbook(export_to_bytes(facade))

    assert set(facade.entities) <= set(workbook.sheetnames)
    inv_sheet = workbook["Investigation"]
    header = [cell.value for cell in inv_sheet[1]]
    assert header == getattr(facade, "Investigation").all_fields
    row = dict(zip(header, [cell.value for cell in inv_sheet[2]], strict=True))
    assert row["unique_id"] == "inv-1"
    assert row["title"] == "Trial"
    study_sheet = workbook["Study"]
    study_header = [cell.value for cell in study_sheet[1]]
    study_row = dict(zip(study_header, [cell.value for cell in study_sheet[2]], strict=True))
    assert study_row["unique_id"] == "st-1"


def test_export_includes_entities_embedded_in_nested_list_fields() -> None:
    client = MetaseedClient("miappe", "1.1")
    client.create_entity(
        "Investigation",
        {
            "unique_id": "inv-2",
            "title": "Nested",
            "contacts": [{"name": "Ada Lovelace", "email": "ada@example.org"}],
        },
        skip_validation=True,
    )

    workbook = load_workbook(export_to_bytes(client.facade))

    person_sheet = workbook["Person"]
    header = [cell.value for cell in person_sheet[1]]
    row = dict(zip(header, [cell.value for cell in person_sheet[2]], strict=True))
    assert row["name"] == "Ada Lovelace"


def test_export_neutralizes_formula_injection() -> None:
    client = MetaseedClient("miappe", "1.1")
    client.create_entity(
        "Investigation",
        {"unique_id": "inv-3", "title": '=HYPERLINK("http://evil")'},
        skip_validation=True,
    )

    workbook = load_workbook(export_to_bytes(client.facade))

    sheet = workbook["Investigation"]
    header = [cell.value for cell in sheet[1]]
    row = dict(zip(header, [cell.value for cell in sheet[2]], strict=True))
    assert row["title"] == '\'=HYPERLINK("http://evil")'


def test_generate_filename_uses_date_profile_version_and_root_id() -> None:
    facade = _facade_with_tree()

    filename = generate_filename(facade)

    date_str = datetime.now().strftime("%y%m%d")
    assert filename == f"{date_str}-miappe-1-1-inv-1.xlsx"


def test_generate_filename_without_entities_falls_back_to_export() -> None:
    facade = MetaseedClient("miappe", "1.1").facade

    filename = generate_filename(facade)

    assert filename.endswith("-export.xlsx")
