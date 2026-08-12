"""Bounded reading and parsing of uploaded files."""

from io import BytesIO
from typing import Any

from fastapi import HTTPException, UploadFile

#: Sheets the export writes for its own purposes rather than for data. Matching
#: on the prefix rather than one name so a later addition needs no change here.
RESERVED_SHEET_PREFIX = "metaseed "

# Upper bound on an uploaded import file, enforced in the app rather than relying
# on the reverse proxy alone. A bare ``await file.read()`` loads the whole upload
# into memory (then parses it), so an unbounded read is a memory-pressure DoS on a
# deployment without an edge cap.
MAX_UPLOAD_BYTES = 25 * 1024 * 1024  # 25 MiB


async def read_upload_capped(file: UploadFile, max_bytes: int = MAX_UPLOAD_BYTES) -> bytes:
    """Read an uploaded file fully, but refuse anything larger than ``max_bytes``.

    Args:
        file: The uploaded file.
        max_bytes: Maximum number of bytes to accept.

    Returns:
        The file content.

    Raises:
        HTTPException: 413 if the upload exceeds ``max_bytes``.
    """
    # Read one byte past the limit to distinguish "exactly at limit" from "over".
    content = await file.read(max_bytes + 1)
    if len(content) > max_bytes:
        raise HTTPException(
            status_code=413,
            detail=f"File too large; the limit is {max_bytes // (1024 * 1024)} MiB.",
        )
    return content


def parse_workbook_sheets(content: bytes) -> dict[str, list[dict[str, Any]]]:
    """Parse an ``.xlsx`` workbook into ``{sheet_name: [row_dicts]}``.

    Each sheet is one entity type, the first row is the header, and template
    placeholder cells/rows (``<field>``) are skipped. Shared by the create-import
    and add-to-existing import routes so the parsing lives in one place.

    Args:
        content: The raw workbook bytes.

    Returns:
        Mapping of sheet name to a list of row dicts (header -> cell value).
    """
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)

    entities_by_type: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith(RESERVED_SHEET_PREFIX):
            # The export's own sheets — the vocabularies behind the dropdowns —
            # are not entities, and reading them back reports every one as an
            # unknown entity type.
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

        for row in rows[1:]:
            # A formatted-but-empty row comes back as an empty tuple in
            # read_only mode, so guard the index before the truthiness check.
            if not row:
                continue

            # Skip placeholder rows (first cell like ``<value>``).
            first_val = str(row[0]) if row[0] else ""
            if first_val.startswith("<") and first_val.endswith(">"):
                continue

            entity_data: dict[str, Any] = {}
            for i, val in enumerate(row):
                if i < len(headers) and val is not None:
                    str_val = str(val)
                    if str_val.startswith("<") and str_val.endswith(">"):
                        continue
                    entity_data[headers[i]] = val

            if entity_data:
                entities_by_type.setdefault(sheet_name, []).append(entity_data)

    return entities_by_type
