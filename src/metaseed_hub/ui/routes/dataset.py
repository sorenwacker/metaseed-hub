"""Dataset routes for Hub UI."""

import copy
import logging
from pathlib import Path
from typing import Annotated, Any

from fastapi import APIRouter, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from fastapi.templating import Jinja2Templates
from metaseed.ui.state import AppState
from sqlalchemy import delete, select
from sqlalchemy.orm import selectinload

from metaseed_hub.models import (
    ChatMessage,
    Comment,
    CommentReaction,
    Dataset,
    DatasetMember,
    DatasetRole,
    DatasetVersion,
    Note,
    ReactionType,
    SpecDraft,
    SpecDraftMember,
    User,
    Workspace,
)
from metaseed_hub.ui.dependencies import (
    CurrentUser,
    DbSession,
    get_dataset_for_user,
    verify_workspace_access,
)
from metaseed_hub.ui.helpers import (
    create_nested_nodes,
    dataset_states,
    ensure_dataset_facade,
    get_dataset_state,
    get_tree_data_from_nodes,
    serialize_tree,
)
from metaseed_hub.ui.render import init_templates as _init_render_templates
from metaseed_hub.ui.render import render_template
from metaseed_hub.ui.security import csrf_error_response, validate_csrf_or_error

logger = logging.getLogger("metaseed_hub")

router = APIRouter(prefix="/datasets", tags=["datasets"])


def init_templates(templates: Jinja2Templates) -> None:
    """Initialize templates reference."""
    _init_render_templates(templates)


@router.get("/new", response_class=HTMLResponse)
async def dataset_new(
    request: Request,
    session: DbSession,
    user: CurrentUser,
    workspace_id: str | None = None,
) -> Response:
    """Return dataset creation form."""
    from metaseed.specs.loader import SpecLoader

    # If no workspace_id provided, get user's first workspace
    if not workspace_id:
        # Get user's tenant using sub (keycloak ID)
        tenant_result = await session.execute(
            select(Workspace)
            .join(Workspace.tenant)
            .where(Workspace.tenant.has(slug=user.sub[:8]))
            .limit(1)
        )
        workspace = tenant_result.scalar_one_or_none()
        if workspace:
            workspace_id = workspace.id
        else:
            # No workspace found - this shouldn't happen normally
            return Response(
                content="<p>No workspace found. Please create one first.</p>", status_code=400
            )

    # Get available profiles and versions from metaseed
    loader = SpecLoader()
    profiles_data = []
    for profile_name in loader.list_profiles():
        versions = loader.list_versions(profile_name)

        # Sort versions in descending order (newest first)
        def version_key(v: str) -> tuple[int, ...]:
            try:
                return tuple(int(x) for x in v.split("."))
            except ValueError:
                return (0,)

        versions = sorted(versions, key=version_key, reverse=True)
        # Get profile metadata from latest version
        display_name = profile_name
        description = ""
        root_entity = "Investigation"
        if versions:
            try:
                spec = loader.load_profile(versions[0], profile_name)
                display_name = spec.display_name or profile_name
                description = spec.description or ""
                root_entity = spec.root_entity or "Investigation"
            except Exception:
                pass
        profiles_data.append(
            {
                "name": profile_name,
                "display_name": display_name,
                "description": description,
                "root_entity": root_entity,
                "versions": versions,
                "latest_version": versions[0] if versions else "",
                "source": "builtin",
            }
        )

    # Get spec drafts from this workspace
    drafts_result = await session.execute(
        select(SpecDraft).where(SpecDraft.workspace_id == workspace_id)
    )
    owned_drafts = list(drafts_result.scalars().all())

    # Also get specs shared with this user via SpecDraftMember
    db_user_result = await session.execute(select(User).where(User.keycloak_id == user.keycloak_id))
    db_user = db_user_result.scalar_one_or_none()

    shared_drafts: list[SpecDraft] = []
    if db_user:
        shared_result = await session.execute(
            select(SpecDraft)
            .join(SpecDraftMember, SpecDraftMember.spec_draft_id == SpecDraft.id)
            .where(SpecDraftMember.user_id == db_user.id)
        )
        shared_drafts = list(shared_result.scalars().all())

    # Combine and deduplicate
    seen_ids: set[str] = set()
    drafts: list[SpecDraft] = []
    for draft in owned_drafts + shared_drafts:
        if draft.id not in seen_ids:
            seen_ids.add(draft.id)
            drafts.append(draft)

    for draft in drafts:
        if draft.name:
            spec_data = draft.spec_data or {}
            profiles_data.append(
                {
                    "name": f"draft:{draft.id}",
                    "display_name": f"{draft.name} (Draft)",
                    "description": spec_data.get("description", ""),
                    "root_entity": spec_data.get("root_entity", "Investigation"),
                    "versions": [draft.version],
                    "latest_version": draft.version,
                    "source": "draft",
                }
            )

    return render_template(
        request=request,
        name="dataset_new.html",
        context={
            "user": user,
            "workspace_id": workspace_id,
            "profiles": profiles_data,
            "user_specs": [],  # TODO: load user's custom specs
            "nav_active": "home",
        },
    )


@router.post("/import")
async def dataset_import(
    request: Request,
    session: DbSession,
    user: CurrentUser,
    file: Annotated[UploadFile, File()],
    name: Annotated[str, Form()],
    workspace_id: Annotated[str | None, Form()] = None,
    profile: Annotated[str | None, Form()] = None,
    version: Annotated[str | None, Form()] = None,
    csrf_token: Annotated[str | None, Form(alias="_csrf_token")] = None,
) -> RedirectResponse:
    """Import a dataset from an uploaded file (JSON, YAML, or Excel)."""
    import json

    import yaml
    from metaseed.models import get_model
    from metaseed.specs.loader import SpecLoader

    from metaseed_hub.ui.helpers import validate_csrf_token

    if not validate_csrf_token(request, csrf_token):
        return RedirectResponse("/hub/?error=csrf_validation_failed", status_code=302)

    # Get workspace if not provided
    if not workspace_id:
        tenant_result = await session.execute(
            select(Workspace)
            .join(Workspace.tenant)
            .where(Workspace.tenant.has(slug=user.sub[:8]))
            .limit(1)
        )
        workspace = tenant_result.scalar_one_or_none()
        if workspace:
            workspace_id = workspace.id
        else:
            return RedirectResponse("/hub/?error=no_workspace", status_code=302)

    await verify_workspace_access(workspace_id, session, user)

    # Read file content
    content = await file.read()
    filename = file.filename or ""

    # Parse based on file type
    data = None
    profile = None
    version = None

    try:
        if filename.endswith((".yaml", ".yml")):
            data = yaml.safe_load(content.decode("utf-8"))
        elif filename.endswith(".json"):
            data = json.loads(content.decode("utf-8"))
        elif filename.endswith((".xlsx", ".xls")):
            # Excel import - read sheets as entity data
            # Each sheet name = entity type, headers = field names
            from io import BytesIO

            import openpyxl

            wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)

            # Parse all sheets into entities dict
            entities_by_type: dict[str, list[dict[str, Any]]] = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    continue

                headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

                # Parse data rows (skip header and placeholder row if it has <field> format)
                for row in rows[1:]:
                    # Skip placeholder rows
                    first_val = str(row[0]) if row[0] else ""
                    if first_val.startswith("<") and first_val.endswith(">"):
                        continue

                    entity_data = {}
                    for i, val in enumerate(row):
                        if i < len(headers) and val is not None:
                            # Skip placeholder values
                            str_val = str(val)
                            if str_val.startswith("<") and str_val.endswith(">"):
                                continue
                            entity_data[headers[i]] = val

                    if entity_data:
                        if sheet_name not in entities_by_type:
                            entities_by_type[sheet_name] = []
                        entities_by_type[sheet_name].append(entity_data)

            # Store entities for later processing
            data = {"_entities_by_type": entities_by_type}

            # Use first entity of first sheet for root if available
            if entities_by_type:
                first_type = list(entities_by_type.keys())[0]
                if entities_by_type[first_type]:
                    data.update(entities_by_type[first_type][0])
        else:
            return RedirectResponse("/hub/datasets/new?error=unsupported_format", status_code=302)

        if not data:
            return RedirectResponse("/hub/datasets/new?error=empty_file", status_code=302)

        # Use provided profile/version, or try to detect from data
        if not profile and isinstance(data, dict):
            profile = data.get("profile") or data.get("_profile")
        if not version and isinstance(data, dict):
            version = data.get("version") or data.get("_version")

        # Default to miappe if not detected
        if not profile:
            profile = "miappe"
        if not version:
            loader = SpecLoader()
            versions = loader.list_versions(profile)
            version = versions[0] if versions else "1.1"

    except Exception as e:
        logger.exception(f"Failed to parse import file: {e}")
        return RedirectResponse("/hub/datasets/new?error=parse_error", status_code=302)

    # Create dataset
    dataset = Dataset(
        workspace_id=workspace_id,
        name=name,
        profile=profile,
        version=version,
        data={},
    )
    session.add(dataset)
    await session.commit()
    await session.refresh(dataset)

    # Try to import entities from data
    try:
        loader = SpecLoader(profile=profile)
        spec = loader.load_profile(version, profile)
        root_entity = spec.root_entity or "Investigation"

        state = get_dataset_state(dataset, dataset_states)
        state.reset()
        state.profile = profile
        state.version = version
        state.facade = None
        facade = state.get_or_create_facade()

        # Handle different data structures
        entities_by_type = data.get("_entities_by_type", {}) if isinstance(data, dict) else {}

        if entities_by_type:
            # Excel import - create entities by type
            # Process root entity first, then children in hierarchy order
            entity_order = [root_entity] + [e for e in facade.entities if e != root_entity]

            for entity_type in entity_order:
                if entity_type not in entities_by_type:
                    continue

                for entity_data in entities_by_type[entity_type]:
                    try:
                        Model = get_model(entity_type, version, profile=profile)
                        # Filter out empty values
                        clean_data = {
                            k: v for k, v in entity_data.items() if v is not None and str(v).strip()
                        }
                        if clean_data:
                            instance = Model(**clean_data)
                            node = state.add_node(entity_type, instance)
                            if not state.editing_node_id:
                                state.editing_node_id = node.id
                    except Exception as e:
                        logger.warning(f"Failed to create {entity_type}: {e}")

        else:
            # JSON/YAML import - use existing logic
            entities_data = data.get("entities", []) if isinstance(data, dict) else []
            if not entities_data and isinstance(data, dict):
                # Try to use the data directly as root entity
                Model = get_model(root_entity, version, profile=profile)
                # Filter out metadata fields
                entity_data = {
                    k: v
                    for k, v in data.items()
                    if not k.startswith("_") and k not in ("profile", "version")
                }
                if entity_data:
                    instance = Model(**entity_data)
                    node = state.add_node(root_entity, instance)
                    state.editing_node_id = node.id
                    create_nested_nodes(
                        state, facade, node, root_entity, copy.deepcopy(entity_data)
                    )

        # Save to database
        from sqlalchemy.orm.attributes import flag_modified

        dataset.data = serialize_tree(state)
        flag_modified(dataset, "data")
        session.add(dataset)
        await session.commit()

    except Exception as e:
        logger.warning(f"Could not import entities, dataset created empty: {e}")

    return RedirectResponse(f"/hub/datasets/{dataset.id}", status_code=303)


@router.post("")
async def dataset_create(
    request: Request,
    session: DbSession,
    user: CurrentUser,
    workspace_id: Annotated[str, Form()],
    name: Annotated[str, Form()],
    profile: Annotated[str, Form()],
    version: Annotated[str, Form()],
    csrf_token: Annotated[str | None, Form(alias="_csrf_token")] = None,
    load_example: Annotated[str | None, Form()] = None,
) -> RedirectResponse:
    """Create a new dataset."""
    import metaseed
    import yaml
    from metaseed.models import get_model
    from metaseed.specs.loader import SpecLoader

    from metaseed_hub.ui.helpers import validate_csrf_token

    if not validate_csrf_token(request, csrf_token):
        url = f"/hub/workspaces/{workspace_id}?error=csrf_validation_failed"
        return RedirectResponse(url, status_code=302)

    # Verify user has access to the workspace
    await verify_workspace_access(workspace_id, session, user)

    # Check if using a draft spec
    spec_draft_id = None
    if profile.startswith("draft:"):
        spec_draft_id = profile.replace("draft:", "")
        # Get the draft to use its name as the profile
        draft_result = await session.execute(select(SpecDraft).where(SpecDraft.id == spec_draft_id))
        draft = draft_result.scalar_one_or_none()
        if draft:
            profile = draft.name
            version = draft.version

    dataset = Dataset(
        workspace_id=workspace_id,
        name=name,
        profile=profile,
        version=version,
        spec_draft_id=spec_draft_id,
        data={},
    )
    session.add(dataset)
    await session.commit()
    await session.refresh(dataset)

    # Load example data if requested
    logger.info(
        f"dataset_create: load_example={load_example!r}, profile={profile}, version={version}"
    )
    if load_example == "true":
        examples_dir = Path(metaseed.__file__).parent / "examples"
        version_dir = examples_dir / profile / version
        yaml_files = list(version_dir.glob("*.yaml")) if version_dir.exists() else []

        if yaml_files:
            try:
                example_data = yaml.safe_load(yaml_files[0].read_text(encoding="utf-8"))
                # Deep copy to prevent Pydantic from modifying the original dict
                example_data_copy = copy.deepcopy(example_data)

                loader = SpecLoader(profile=profile)
                spec = loader.load_profile(version, profile)
                root_entity = spec.root_entity or "Investigation"

                state = get_dataset_state(dataset, dataset_states)
                state.reset()
                state.profile = profile
                state.version = version
                state.facade = None
                facade = state.get_or_create_facade()

                Model = get_model(root_entity, version, profile=profile)
                instance = Model(**example_data)
                node = state.add_node(root_entity, instance)
                state.editing_node_id = node.id

                # Create nested child nodes from the unmodified copy
                create_nested_nodes(state, facade, node, root_entity, example_data_copy)

                from sqlalchemy.orm.attributes import flag_modified

                dataset.data = serialize_tree(state)
                flag_modified(dataset, "data")
                session.add(dataset)
                await session.commit()
            except Exception as e:
                logger.exception(f"Failed to load example data: {e}")

    return RedirectResponse(f"/hub/datasets/{dataset.id}", status_code=303)


@router.delete("/{dataset_id}", response_class=HTMLResponse)
async def dataset_delete(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Delete a dataset."""
    try:
        validate_csrf_or_error(request)
    except Exception:
        return csrf_error_response()

    try:
        # Verify user has access to this dataset
        dataset = await get_dataset_for_user(dataset_id, session, user)
    except Exception as e:
        logger.error(f"Failed to get dataset {dataset_id}: {e}")
        response = Response(status_code=200)
        response.headers["HX-Trigger"] = (
            '{"showToast": {"message": "Dataset not found or access denied", "type": "error"}}'
        )
        return response

    workspace_id = dataset.workspace_id

    # Clear state cache
    if dataset_id in dataset_states:
        del dataset_states[dataset_id]

    try:
        # Manually delete related records (in case CASCADE not set in DB)
        await session.execute(
            delete(CommentReaction).where(
                CommentReaction.comment_id.in_(
                    select(Comment.id).where(Comment.dataset_id == dataset_id)
                )
            )
        )
        await session.execute(delete(Comment).where(Comment.dataset_id == dataset_id))
        await session.execute(delete(ChatMessage).where(ChatMessage.dataset_id == dataset_id))
        await session.execute(delete(Note).where(Note.dataset_id == dataset_id))
        await session.execute(delete(DatasetMember).where(DatasetMember.dataset_id == dataset_id))
        await session.execute(delete(DatasetVersion).where(DatasetVersion.dataset_id == dataset_id))

        await session.delete(dataset)
        await session.commit()
    except Exception as e:
        logger.error(f"Failed to delete dataset {dataset_id}: {e}", exc_info=True)
        await session.rollback()
        error_msg = str(e).replace('"', '\\"')
        response = Response(status_code=200)
        response.headers["HX-Trigger"] = (
            f'{{"showToast": {{"message": "Delete failed: {error_msg}", "type": "error"}}}}'
        )
        return response

    # If request target is body, redirect to home (called from dataset page)
    hx_target = request.headers.get("HX-Target", "")
    if hx_target == "body":
        # Return HTML with redirect script as fallback
        response = Response(
            content='<script>window.location.href="/hub/";</script>',
            status_code=200,
            media_type="text/html",
        )
        response.headers["HX-Redirect"] = "/hub/"
        return response

    # Return updated dataset grid
    result = await session.execute(select(Dataset).where(Dataset.workspace_id == workspace_id))
    datasets = list(result.scalars().all())

    return render_template(
        request=request,
        name="partials/dataset_grid.html",
        context={"datasets": datasets},
    )


@router.post("/{dataset_id}/load-example", response_class=HTMLResponse)
async def dataset_load_example(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Load example data into a dataset from YAML files."""
    import metaseed
    import yaml
    from metaseed.models import get_model
    from metaseed.specs.loader import SpecLoader

    try:
        validate_csrf_or_error(request)
    except Exception:
        return csrf_error_response()

    # Verify user has access to this dataset
    dataset = await get_dataset_for_user(dataset_id, session, user)

    # Find example YAML file
    examples_dir = Path(metaseed.__file__).parent / "examples"
    version_dir = examples_dir / dataset.profile / dataset.version

    if not version_dir.exists():
        msg = f"No example available for {dataset.profile} v{dataset.version}"
        return HTMLResponse(f"<div class='error'>{msg}</div>")

    yaml_files = list(version_dir.glob("*.yaml"))
    if not yaml_files:
        msg = f"No example file found for {dataset.profile} v{dataset.version}"
        return HTMLResponse(f"<div class='error'>{msg}</div>")

    example_file = yaml_files[0]
    try:
        example_data = yaml.safe_load(example_file.read_text(encoding="utf-8"))
    except yaml.YAMLError as e:
        return HTMLResponse(f"<div class='error'>Error loading example: {e}</div>")

    # Deep copy to prevent Pydantic from modifying the original dict
    example_data_copy = copy.deepcopy(example_data)

    # Load spec to get root entity
    loader = SpecLoader(profile=dataset.profile)
    spec = loader.load_profile(dataset.version, dataset.profile)
    root_entity = spec.root_entity or "Investigation"

    # Load example into dataset state (append, don't replace)
    state = get_dataset_state(dataset, dataset_states)
    facade = state.get_or_create_facade()

    try:
        Model = get_model(root_entity, dataset.version, profile=dataset.profile)
        instance = Model(**example_data)
        node = state.add_node(root_entity, instance)
        state.editing_node_id = node.id

        # Create nested child nodes from the unmodified copy
        create_nested_nodes(state, facade, node, root_entity, example_data_copy)

        # Save to database
        from sqlalchemy.orm.attributes import flag_modified

        dataset.data = serialize_tree(state)
        logger.info(f"Saving tree with {len(state.entity_tree)} root nodes")
        for n in state.entity_tree:
            if n.instance:
                data = n.instance.model_dump(exclude_none=True)
                logger.info(f"  {n.entity_type} '{n.label}': {len(data)} fields")
            for c in n.children[:3]:
                if c.instance:
                    cdata = c.instance.model_dump(exclude_none=True)
                    logger.info(f"    Child {c.entity_type} '{c.label}': {len(cdata)} fields")

        flag_modified(dataset, "data")
        session.add(dataset)
        await session.commit()

    except Exception as e:
        import traceback

        logger.exception(f"Failed to load example: {e}")
        tb = traceback.format_exc()
        error_html = f"""
        <div class='notification error' style='user-select: text;'>
            <strong>Error loading example:</strong>
            <pre style='white-space: pre-wrap; font-size: 0.75rem; margin-top: 0.5rem;'>{e}

{tb}</pre>
        </div>"""
        return HTMLResponse(error_html)

    # Use HX-Redirect for HTMX to do a full page redirect
    response = HTMLResponse(status_code=200)
    response.headers["HX-Redirect"] = f"/hub/datasets/{dataset_id}"
    return response


def _get_entity_info(state: AppState) -> list[dict[str, str]]:
    """Get entity type information including descriptions."""
    facade = state.get_or_create_facade()
    entity_info = []
    for entity_name in facade.entities:
        helper = getattr(facade, entity_name, None)
        if helper:
            entity_info.append(
                {
                    "name": entity_name,
                    "description": helper.description or "",
                }
            )
    return entity_info


async def _build_dataset_context(
    dataset: Dataset,
    session: Any,
) -> dict[str, Any]:
    """Build common context for dataset views.

    Args:
        dataset: Dataset model.
        session: Database session for loading spec drafts.

    Returns:
        Dictionary with state, tree_data, and entity_descriptions.
    """
    state = await ensure_dataset_facade(dataset, session)
    tree_data = get_tree_data_from_nodes(state)

    # Get descriptions for entity types
    entity_descriptions: dict[str, str] = {}

    try:
        facade = state.get_or_create_facade()
        for entity_name in facade.entities:
            helper = getattr(facade, entity_name, None)
            if helper:
                entity_descriptions[entity_name] = helper.description or ""

    except Exception as e:
        # Log but don't crash
        logger.warning(f"Failed to load facade for dataset {dataset.id}: {e}")

    return {
        "state": state,
        "tree_data": tree_data,
        "entity_descriptions": entity_descriptions,
    }


@router.get("/{dataset_id}", response_class=HTMLResponse)
async def dataset_editor(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Dataset editor - wraps metaseed UI with hub chrome."""
    # Verify user has access to this dataset
    dataset = await get_dataset_for_user(dataset_id, session, user)

    # Load workspace for breadcrumb
    workspace = await session.get(Workspace, dataset.workspace_id)

    # Load members with user info
    members_result = await session.execute(
        select(DatasetMember)
        .where(DatasetMember.dataset_id == dataset_id)
        .options(selectinload(DatasetMember.user))
    )
    members = list(members_result.scalars().all())

    # Build common dataset context
    try:
        ctx = await _build_dataset_context(dataset, session)
        root_types = ctx["state"].get_root_entity_types()
    except Exception as e:
        # Profile doesn't exist or is invalid - show error page
        logger.warning(f"Failed to load dataset {dataset_id}: {e}")
        return render_template(
            request=request,
            name="dataset.html",
            context={
                "user": user,
                "dataset": dataset,
                "workspace": workspace,
                "members": members,
                "state": None,
                "root_types": [],
                "tree_data": [],
                "entity_descriptions": {},
                "nav_active": "home",
                "error": (
                    f"Could not load profile '{dataset.profile}' v{dataset.version}. "
                    "The profile may not exist or may be invalid."
                ),
            },
        )

    return render_template(
        request=request,
        name="dataset.html",
        context={
            "user": user,
            "dataset": dataset,
            "workspace": workspace,
            "members": members,
            "state": ctx["state"],
            "root_types": root_types,
            "tree_data": ctx["tree_data"],
            "entity_descriptions": ctx["entity_descriptions"],
            "nav_active": "home",
        },
    )


@router.get("/{dataset_id}/tree", response_class=HTMLResponse)
async def dataset_tree(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Return entity tree for a dataset."""
    # Verify user has access to this dataset
    dataset = await get_dataset_for_user(dataset_id, session, user)

    state = await ensure_dataset_facade(dataset, session)
    tree_data = get_tree_data_from_nodes(state)

    return render_template(
        request=request,
        name="partials/entity_tree.html",
        context={
            "tree_data": tree_data,
            "dataset_id": dataset_id,
        },
    )


@router.get("/{dataset_id}/root-buttons", response_class=HTMLResponse)
async def dataset_root_buttons(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Return root entity type buttons for the sidebar."""
    dataset = await get_dataset_for_user(dataset_id, session, user)

    state = await ensure_dataset_facade(dataset, session)
    root_types = state.get_root_entity_types()
    tree_data = get_tree_data_from_nodes(state)
    existing_root_types = [item["entity_type"] for item in tree_data]

    return render_template(
        request=request,
        name="partials/root_buttons.html",
        context={
            "dataset_id": dataset_id,
            "root_types": root_types,
            "existing_root_types": existing_root_types,
        },
    )


@router.get("/{dataset_id}/overview", response_class=HTMLResponse)
async def dataset_overview(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Return the overview panel with version history, comments, and sharing."""
    dataset = await get_dataset_for_user(dataset_id, session, user)

    # Load members
    members_result = await session.execute(
        select(DatasetMember)
        .where(DatasetMember.dataset_id == dataset_id)
        .options(selectinload(DatasetMember.user))
    )
    members = list(members_result.scalars().all())

    # Check if there's any tree data
    state = await ensure_dataset_facade(dataset, session)
    tree_data = get_tree_data_from_nodes(state)

    return render_template(
        request=request,
        name="partials/dataset_overview.html",
        context={
            "dataset": dataset,
            "members": members,
            "tree_data": tree_data,
        },
    )


@router.post("/{dataset_id}/validate", response_class=HTMLResponse)
async def dataset_validate(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
) -> HTMLResponse:
    """Validate all entities in the dataset against their schemas."""
    from pydantic import ValidationError

    try:
        validate_csrf_or_error(request)
    except Exception:
        return csrf_error_response()

    # Verify user has access to this dataset
    dataset = await get_dataset_for_user(dataset_id, session, user)

    state = await ensure_dataset_facade(dataset, session)
    facade = state.get_or_create_facade()

    errors: list[dict[str, Any]] = []

    for node_id, node in state.nodes_by_id.items():
        try:
            helper = getattr(facade, node.entity_type)
            data = node.instance.model_dump(exclude_none=True) if node.instance else {}
            # Re-validate by recreating - Pydantic validation runs here
            helper.create(**data)
        except ValidationError as e:
            errors.append(
                {
                    "node_id": node_id,
                    "entity_type": node.entity_type,
                    "label": node.label,
                    "errors": [
                        {"field": ".".join(str(x) for x in err["loc"]), "message": err["msg"]}
                        for err in e.errors()
                    ],
                }
            )
        except AttributeError:
            errors.append(
                {
                    "node_id": node_id,
                    "entity_type": node.entity_type,
                    "label": node.label,
                    "errors": [
                        {"field": "", "message": f"Unknown entity type: {node.entity_type}"}
                    ],
                }
            )
        except Exception as e:
            errors.append(
                {
                    "node_id": node_id,
                    "entity_type": node.entity_type,
                    "label": node.label,
                    "errors": [{"field": "", "message": str(e)}],
                }
            )

    # Build HTML response
    total = len(state.nodes_by_id)
    if not errors:
        html = f"""
        <div class="validation-results success">
            <div class="validation-header success-message">
                All {total} entities are valid.
            </div>
        </div>
        """
    else:
        html = f"""
        <div class="validation-results">
            <div class="validation-header error-message">
                {len(errors)} of {total} entities have validation errors.
            </div>
            <div class="validation-errors">
        """
        for err in errors:
            html += f"""
            <div class="validation-error-item">
                <div class="validation-entity">
                    <span class="entity-type-badge">{err["entity_type"]}</span>
                    <a href="#" class="entity-link"
                       hx-get="/hub/datasets/{dataset_id}/entity/{err["node_id"]}"
                       hx-target="#editor"
                       hx-swap="innerHTML">{err["label"]}</a>
                </div>
                <ul class="validation-error-list">
            """
            for field_err in err["errors"]:
                field = field_err["field"] or "(general)"
                html += f"<li><strong>{field}:</strong> {field_err['message']}</li>"
            html += "</ul></div>"
        html += "</div></div>"

    return HTMLResponse(html)


@router.get("/{dataset_id}/graph", response_class=HTMLResponse)
async def dataset_graph(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
    node_id: str | None = None,
) -> Response:
    """Graph visualization of dataset entities.

    Args:
        node_id: Optional. If provided, only show this entity and its descendants.
    """
    # Verify user has access to this dataset
    dataset = await get_dataset_for_user(dataset_id, session, user)

    return render_template(
        request=request,
        name="graph.html",
        context={
            "user": user,
            "dataset": dataset,
            "node_id": node_id,
            "nav_active": "home",
        },
    )


@router.get("/{dataset_id}/api/graph")
async def dataset_graph_api(
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
    node_id: str | None = None,
) -> Response:
    """Return graph data for visualization (JSON API).

    Uses metaseed's build_graph to extract nodes and edges from instance data.

    Args:
        node_id: Optional. If provided, only include this node and its descendants.
    """
    from fastapi.responses import JSONResponse
    from metaseed.ui.services.graph import build_graph

    # Verify user has access to this dataset
    dataset = await get_dataset_for_user(dataset_id, session, user)

    state = await ensure_dataset_facade(dataset, session)

    # Use metaseed's graph builder which properly extracts nested entities
    graph_data = build_graph(state)

    return JSONResponse(content=graph_data)


@router.get("/{dataset_id}/chat", response_class=HTMLResponse)
async def dataset_chat_page(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Full-page chat view for a dataset."""
    dataset = await get_dataset_for_user(dataset_id, session, user)

    return render_template(
        request=request,
        name="chat.html",
        context={
            "user": user,
            "dataset": dataset,
            "nav_active": "home",
        },
    )


@router.post("/{dataset_id}/chat", response_class=HTMLResponse)
async def dataset_chat(
    request: Request,
    dataset_id: str,
    user: CurrentUser,
    message: Annotated[str, Form()],
) -> HTMLResponse:
    """Post a chat message."""
    import html as html_module

    try:
        validate_csrf_or_error(request)
    except Exception:
        return csrf_error_response()

    # Escape user content to prevent XSS
    safe_name = html_module.escape(user.name or "")
    safe_message = html_module.escape(message)
    html = f"<div class='chat-message'><strong>{safe_name}:</strong> {safe_message}</div>"
    return HTMLResponse(html)


@router.get("/{dataset_id}/export")
async def dataset_export(
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Export dataset data to Excel file.

    Uses metaseed's export service to generate an Excel workbook
    containing all entities in the dataset.
    """
    from fastapi.responses import StreamingResponse
    from metaseed.ui.services.export import export_to_bytes, generate_filename

    dataset = await get_dataset_for_user(dataset_id, session, user)
    state = await ensure_dataset_facade(dataset, session)

    # Generate Excel file using metaseed's export service
    excel_bytes = export_to_bytes(state)
    filename = generate_filename(state)

    # If no data, use dataset name for filename
    if not filename or filename == "export.xlsx":
        filename = f"{dataset.name.replace(' ', '_')}.xlsx"

    return StreamingResponse(
        excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/{dataset_id}/import")
async def dataset_import_into_existing(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
    file: Annotated[UploadFile, File()],
) -> Response:
    """Import data from file into an existing dataset.

    Supports JSON, YAML, and Excel files. Adds entities to the existing dataset.
    """
    import json
    from io import BytesIO

    import yaml
    from metaseed.models import get_model
    from sqlalchemy.orm.attributes import flag_modified

    try:
        validate_csrf_or_error(request)
    except Exception:
        return csrf_error_response()

    dataset = await get_dataset_for_user(dataset_id, session, user)

    # Read file content
    content = await file.read()
    filename = file.filename or ""

    # Parse based on file type
    entities_by_type: dict[str, list[dict[str, Any]]] = {}

    try:
        if filename.endswith((".yaml", ".yml")):
            data = yaml.safe_load(content.decode("utf-8"))
            # For YAML, treat as single root entity or check for entities list
            if isinstance(data, dict) and "entities" in data:
                for entity in data["entities"]:
                    etype = entity.get("_type", dataset.profile)
                    if etype not in entities_by_type:
                        entities_by_type[etype] = []
                    entities_by_type[etype].append(entity)
            elif isinstance(data, dict):
                # Single entity - use root entity type
                from metaseed.specs.loader import SpecLoader

                loader = SpecLoader(profile=dataset.profile)
                spec = loader.load_profile(dataset.version, dataset.profile)
                root_entity = spec.root_entity or "Investigation"
                entities_by_type[root_entity] = [data]

        elif filename.endswith(".json"):
            data = json.loads(content.decode("utf-8"))
            if isinstance(data, dict) and "entities" in data:
                for entity in data["entities"]:
                    etype = entity.get("_type", dataset.profile)
                    if etype not in entities_by_type:
                        entities_by_type[etype] = []
                    entities_by_type[etype].append(entity)
            elif isinstance(data, dict):
                from metaseed.specs.loader import SpecLoader

                loader = SpecLoader(profile=dataset.profile)
                spec = loader.load_profile(dataset.version, dataset.profile)
                root_entity = spec.root_entity or "Investigation"
                entities_by_type[root_entity] = [data]

        elif filename.endswith((".xlsx", ".xls")):
            import openpyxl

            wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    continue

                headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

                for row in rows[1:]:
                    first_val = str(row[0]) if row[0] else ""
                    if first_val.startswith("<") and first_val.endswith(">"):
                        continue

                    entity_data: dict[str, Any] = {}
                    for i, val in enumerate(row):
                        if i < len(headers) and val is not None:
                            str_val = str(val)
                            if str_val.startswith("<") and str_val.endswith(">"):
                                continue
                            entity_data[headers[i]] = val

                    if entity_data:
                        if sheet_name not in entities_by_type:
                            entities_by_type[sheet_name] = []
                        entities_by_type[sheet_name].append(entity_data)
        else:
            return HTMLResponse(
                "<div class='notification error'>Unsupported file format</div>",
                status_code=400,
            )

    except Exception as e:
        logger.exception(f"Failed to parse import file: {e}")
        return HTMLResponse(
            f"<div class='notification error'>Parse error: {e}</div>",
            status_code=400,
        )

    # Get dataset state and add entities
    state = await ensure_dataset_facade(dataset, session)
    facade = state.get_or_create_facade()

    imported_count = 0
    errors: list[str] = []

    # Process entities in order (root first)
    from metaseed.specs.loader import SpecLoader

    loader = SpecLoader(profile=dataset.profile)
    spec = loader.load_profile(dataset.version, dataset.profile)
    root_entity = spec.root_entity or "Investigation"

    entity_order = [root_entity] + [e for e in facade.entities if e != root_entity]

    for entity_type in entity_order:
        if entity_type not in entities_by_type:
            continue

        for entity_data in entities_by_type[entity_type]:
            try:
                Model = get_model(entity_type, dataset.version, profile=dataset.profile)
                clean_data = {
                    k: v
                    for k, v in entity_data.items()
                    if v is not None and str(v).strip() and not k.startswith("_")
                }
                if clean_data:
                    instance = Model(**clean_data)
                    state.add_node(entity_type, instance)
                    imported_count += 1
            except Exception as e:
                errors.append(f"{entity_type}: {e}")

    # Save to database
    dataset.data = serialize_tree(state)
    flag_modified(dataset, "data")
    session.add(dataset)
    await session.commit()

    if errors:
        msg = f"Imported {imported_count} entities with {len(errors)} errors"
        logger.warning(f"Import errors: {errors[:5]}")
    else:
        msg = f"Successfully imported {imported_count} entities"

    return HTMLResponse(f"<div class='notification success'>{msg}</div>")


# =============================================================================
# Member Management Routes
# =============================================================================


async def _get_members_html(
    request: Request,
    dataset_id: str,
    session: DbSession,
) -> Response:
    """Render the member list partial."""
    result = await session.execute(
        select(DatasetMember)
        .where(DatasetMember.dataset_id == dataset_id)
        .options(selectinload(DatasetMember.user))
    )
    members = list(result.scalars().all())

    return render_template(
        request=request,
        name="partials/dataset_members.html",
        context={
            "members": members,
            "dataset_id": dataset_id,
        },
    )


@router.post("/{dataset_id}/members", response_class=HTMLResponse)
async def add_dataset_member(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
    email: Annotated[str, Form()],
) -> Response:
    """Add a member to a dataset by email."""
    # Verify user has access
    await get_dataset_for_user(dataset_id, session, user)

    # Find user by email
    result = await session.execute(select(User).where(User.email == email))
    target_user = result.scalar_one_or_none()

    if not target_user:
        # Return error message - user must log in first
        response = await _get_members_html(request, dataset_id, session)
        msg = "User not found. They must log in first before you can share."
        response.headers["HX-Trigger"] = f'{{"showToast": {{"message": "{msg}", "type": "error"}}}}'
        return response

    # Check if already a member
    existing = await session.execute(
        select(DatasetMember).where(
            DatasetMember.dataset_id == dataset_id,
            DatasetMember.user_id == target_user.id,
        )
    )
    if existing.scalar_one_or_none():
        return await _get_members_html(request, dataset_id, session)

    # Add member with viewer role by default
    member = DatasetMember(
        dataset_id=dataset_id,
        user_id=target_user.id,
        role=DatasetRole.VIEWER,
    )
    session.add(member)
    await session.commit()

    # Refresh to ensure we can load relationships
    await session.refresh(member)

    return await _get_members_html(request, dataset_id, session)


@router.patch("/{dataset_id}/members/{user_id}", response_class=HTMLResponse)
async def update_dataset_member_role(
    request: Request,
    dataset_id: str,
    user_id: str,
    session: DbSession,
    user: CurrentUser,
    role: Annotated[str, Form()],
) -> Response:
    """Update a member's role in a dataset."""
    # Verify user has access
    await get_dataset_for_user(dataset_id, session, user)

    # Find membership
    result = await session.execute(
        select(DatasetMember).where(
            DatasetMember.dataset_id == dataset_id,
            DatasetMember.user_id == user_id,
        )
    )
    member = result.scalar_one_or_none()

    if member:
        member.role = DatasetRole(role)
        await session.commit()

    return await _get_members_html(request, dataset_id, session)


@router.delete("/{dataset_id}/members/{user_id}", response_class=HTMLResponse)
async def remove_dataset_member(
    request: Request,
    dataset_id: str,
    user_id: str,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Remove a member from a dataset."""
    # Verify user has access
    await get_dataset_for_user(dataset_id, session, user)

    # Find and delete membership
    result = await session.execute(
        select(DatasetMember).where(
            DatasetMember.dataset_id == dataset_id,
            DatasetMember.user_id == user_id,
        )
    )
    member = result.scalar_one_or_none()

    if member:
        await session.delete(member)
        await session.commit()

    return await _get_members_html(request, dataset_id, session)


# =============================================================================
# Comment Routes (Threaded, with reactions)
# =============================================================================


async def _get_comments_html(
    request: Request,
    dataset_id: str,
    session: DbSession,
    keycloak_sub: str,
) -> Response:
    """Render the comments list partial."""
    # Get database user ID from keycloak sub
    user_result = await session.execute(select(User).where(User.keycloak_id == keycloak_sub))
    db_user = user_result.scalar_one_or_none()
    current_user_id = db_user.id if db_user else None

    # Get top-level comments (no parent) with all nested relationships eagerly loaded
    result = await session.execute(
        select(Comment)
        .where(Comment.dataset_id == dataset_id, Comment.parent_id.is_(None))
        .options(
            selectinload(Comment.user),
            selectinload(Comment.reactions),
            selectinload(Comment.replies).selectinload(Comment.user),
            selectinload(Comment.replies).selectinload(Comment.reactions),
            selectinload(Comment.replies).selectinload(Comment.replies).selectinload(Comment.user),
            selectinload(Comment.replies)
            .selectinload(Comment.replies)
            .selectinload(Comment.reactions),
        )
        .order_by(Comment.created_at.desc())
    )
    comments = list(result.scalars().all())

    return render_template(
        request=request,
        name="partials/comments_list.html",
        context={
            "comments": comments,
            "dataset_id": dataset_id,
            "current_user_id": current_user_id,
        },
    )


@router.get("/{dataset_id}/comments", response_class=HTMLResponse)
async def get_dataset_comments(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Get all comments for a dataset."""
    await get_dataset_for_user(dataset_id, session, user)
    return await _get_comments_html(request, dataset_id, session, user.sub)


@router.post("/{dataset_id}/comments", response_class=HTMLResponse)
async def add_dataset_comment(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
    content: Annotated[str, Form()],
    parent_id: Annotated[str | None, Form()] = None,
) -> Response:
    """Add a comment to a dataset."""
    try:
        validate_csrf_or_error(request)
    except Exception:
        return csrf_error_response()

    await get_dataset_for_user(dataset_id, session, user)

    # Get user from database by keycloak sub
    user_result = await session.execute(select(User).where(User.keycloak_id == user.sub))
    db_user = user_result.scalar_one_or_none()

    if not db_user:
        return HTMLResponse("<div class='error'>User not found</div>", status_code=400)

    comment = Comment(
        dataset_id=dataset_id,
        user_id=db_user.id,
        parent_id=parent_id if parent_id else None,
        content=content.strip(),
    )
    session.add(comment)
    await session.commit()

    return await _get_comments_html(request, dataset_id, session, user.sub)


@router.delete("/{dataset_id}/comments/{comment_id}", response_class=HTMLResponse)
async def delete_dataset_comment(
    request: Request,
    dataset_id: str,
    comment_id: str,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Delete a comment (only by owner)."""
    try:
        validate_csrf_or_error(request)
    except Exception:
        return csrf_error_response()

    await get_dataset_for_user(dataset_id, session, user)

    # Get user from database
    user_result = await session.execute(select(User).where(User.keycloak_id == user.sub))
    db_user = user_result.scalar_one_or_none()

    if not db_user:
        return HTMLResponse("<div class='error'>User not found</div>", status_code=400)

    # Find comment and verify ownership
    result = await session.execute(select(Comment).where(Comment.id == comment_id))
    comment = result.scalar_one_or_none()

    if comment and comment.user_id == db_user.id:
        await session.delete(comment)
        await session.commit()

    return await _get_comments_html(request, dataset_id, session, user.sub)


@router.post("/{dataset_id}/comments/{comment_id}/react", response_class=HTMLResponse)
async def react_to_comment(
    request: Request,
    dataset_id: str,
    comment_id: str,
    session: DbSession,
    user: CurrentUser,
    reaction: Annotated[str, Form()],
) -> Response:
    """Add or toggle a reaction on a comment."""
    try:
        validate_csrf_or_error(request)
    except Exception:
        return csrf_error_response()

    await get_dataset_for_user(dataset_id, session, user)

    # Get user from database
    user_result = await session.execute(select(User).where(User.keycloak_id == user.sub))
    db_user = user_result.scalar_one_or_none()

    if not db_user:
        return HTMLResponse("<div class='error'>User not found</div>", status_code=400)

    # Check for existing reaction
    existing_result = await session.execute(
        select(CommentReaction).where(
            CommentReaction.comment_id == comment_id,
            CommentReaction.user_id == db_user.id,
        )
    )
    existing = existing_result.scalar_one_or_none()

    reaction_type = ReactionType(reaction)

    if existing:
        if existing.reaction == reaction_type:
            # Toggle off - remove reaction
            await session.delete(existing)
        else:
            # Change reaction type
            existing.reaction = reaction_type
    else:
        # Add new reaction
        new_reaction = CommentReaction(
            comment_id=comment_id,
            user_id=db_user.id,
            reaction=reaction_type,
        )
        session.add(new_reaction)

    await session.commit()

    return await _get_comments_html(request, dataset_id, session, user.sub)


# =============================================================================
# Version History Routes
# =============================================================================


def _flatten_tree(tree: list[dict[str, Any]], prefix: str = "") -> dict[str, dict[str, Any]]:
    """Flatten tree into dict keyed by node ID with full data."""
    result: dict[str, dict[str, Any]] = {}
    for node in tree:
        node_id = node.get("id", "")
        if node_id:
            result[node_id] = {
                "entity_type": node.get("entity_type", "Unknown"),
                "label": node.get("label", ""),
                "data": node.get("data", {}),
            }
        if "children" in node:
            result.update(_flatten_tree(node["children"], prefix))
    return result


def _calculate_diff(old_data: dict[str, Any], new_data: dict[str, Any]) -> dict[str, Any]:
    """Calculate diff between two dataset states.

    Returns dict with summary changes and detailed field changes.
    """
    old_tree = old_data.get("tree", [])
    new_tree = new_data.get("tree", [])

    def count_entities(tree: list[dict[str, Any]]) -> dict[str, int]:
        """Count entities by type in tree."""
        counts: dict[str, int] = {}
        for node in tree:
            entity_type = node.get("entity_type", "Unknown")
            counts[entity_type] = counts.get(entity_type, 0) + 1
            if "children" in node:
                child_counts = count_entities(node["children"])
                for k, v in child_counts.items():
                    counts[k] = counts.get(k, 0) + v
        return counts

    old_counts = count_entities(old_tree)
    new_counts = count_entities(new_tree)

    all_types = set(old_counts.keys()) | set(new_counts.keys())
    changes: list[str] = []

    for entity_type in sorted(all_types):
        old_count = old_counts.get(entity_type, 0)
        new_count = new_counts.get(entity_type, 0)
        if new_count > old_count:
            changes.append(f"+{new_count - old_count} {entity_type}")
        elif new_count < old_count:
            changes.append(f"-{old_count - new_count} {entity_type}")

    total_old = sum(old_counts.values())
    total_new = sum(new_counts.values())

    return {
        "changes": changes,
        "total_old": total_old,
        "total_new": total_new,
        "has_changes": changes or (old_data != new_data),
    }


def _calculate_detailed_diff(
    old_data: dict[str, Any], new_data: dict[str, Any]
) -> list[dict[str, Any]]:
    """Calculate detailed field-level diff between two dataset states.

    Returns list of changes with entity info and field diffs.
    """
    old_nodes = _flatten_tree(old_data.get("tree", []))
    new_nodes = _flatten_tree(new_data.get("tree", []))

    all_ids = set(old_nodes.keys()) | set(new_nodes.keys())
    changes: list[dict[str, Any]] = []

    for node_id in all_ids:
        old_node = old_nodes.get(node_id)
        new_node = new_nodes.get(node_id)

        if old_node and not new_node:
            # Entity removed
            changes.append(
                {
                    "type": "removed",
                    "entity_type": old_node["entity_type"],
                    "label": old_node["label"],
                    "fields": [],
                }
            )
        elif new_node and not old_node:
            # Entity added
            field_changes = []
            for key, val in new_node["data"].items():
                if val is not None and str(val).strip():
                    field_changes.append(
                        {
                            "field": key,
                            "old": None,
                            "new": val,
                        }
                    )
            changes.append(
                {
                    "type": "added",
                    "entity_type": new_node["entity_type"],
                    "label": new_node["label"],
                    "fields": field_changes,
                }
            )
        elif old_node and new_node:
            # Check for field changes
            old_fields = old_node["data"]
            new_fields = new_node["data"]
            all_keys = set(old_fields.keys()) | set(new_fields.keys())

            field_changes = []
            for key in all_keys:
                old_val = old_fields.get(key)
                new_val = new_fields.get(key)
                if old_val != new_val:
                    field_changes.append(
                        {
                            "field": key,
                            "old": old_val,
                            "new": new_val,
                        }
                    )

            if field_changes:
                changes.append(
                    {
                        "type": "modified",
                        "entity_type": new_node["entity_type"],
                        "label": new_node["label"],
                        "fields": field_changes,
                    }
                )

    return changes


@router.get("/{dataset_id}/versions", response_class=HTMLResponse)
async def get_dataset_versions(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Get version history for a dataset with diffs."""
    await get_dataset_for_user(dataset_id, session, user)

    result = await session.execute(
        select(DatasetVersion)
        .where(DatasetVersion.dataset_id == dataset_id)
        .options(selectinload(DatasetVersion.created_by))
        .order_by(DatasetVersion.version_number.desc())
    )
    versions = list(result.scalars().all())

    # Calculate diffs between consecutive versions
    versions_with_diffs: list[dict[str, Any]] = []
    for i, version in enumerate(versions):
        version_data: dict[str, Any] = {
            "version": version,
            "diff": None,
        }
        # Compare with previous version (next in list since sorted desc)
        if i < len(versions) - 1:
            prev_version = versions[i + 1]
            version_data["diff"] = _calculate_diff(prev_version.data, version.data)
        elif i == len(versions) - 1:
            # First version - compare with empty
            version_data["diff"] = _calculate_diff({}, version.data)

        versions_with_diffs.append(version_data)

    return render_template(
        request=request,
        name="partials/dataset_versions.html",
        context={
            "versions": versions_with_diffs,
            "dataset_id": dataset_id,
        },
    )


@router.get("/{dataset_id}/versions/{version_id}/diff", response_class=HTMLResponse)
async def get_version_diff(
    request: Request,
    dataset_id: str,
    version_id: str,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Get detailed diff for a specific version."""
    await get_dataset_for_user(dataset_id, session, user)

    # Get the version
    result = await session.execute(
        select(DatasetVersion).where(
            DatasetVersion.id == version_id, DatasetVersion.dataset_id == dataset_id
        )
    )
    version = result.scalar_one_or_none()

    if not version:
        return HTMLResponse("<div class='error'>Version not found</div>", status_code=404)

    # Get previous version
    prev_result = await session.execute(
        select(DatasetVersion)
        .where(
            DatasetVersion.dataset_id == dataset_id,
            DatasetVersion.version_number < version.version_number,
        )
        .order_by(DatasetVersion.version_number.desc())
        .limit(1)
    )
    prev_version = prev_result.scalar_one_or_none()

    # Calculate detailed diff
    old_data = prev_version.data if prev_version else {}
    detailed_changes = _calculate_detailed_diff(old_data, version.data)

    return render_template(
        request=request,
        name="partials/version_diff.html",
        context={
            "version": version,
            "changes": detailed_changes,
            "dataset_id": dataset_id,
        },
    )


@router.post("/{dataset_id}/versions/{version_id}/restore", response_class=HTMLResponse)
async def restore_dataset_version(
    request: Request,
    dataset_id: str,
    version_id: str,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Restore a dataset to a previous version."""
    try:
        validate_csrf_or_error(request)
    except Exception:
        return csrf_error_response()

    dataset = await get_dataset_for_user(dataset_id, session, user)

    # Get the version to restore
    result = await session.execute(
        select(DatasetVersion).where(
            DatasetVersion.id == version_id, DatasetVersion.dataset_id == dataset_id
        )
    )
    version = result.scalar_one_or_none()

    if not version:
        return HTMLResponse(
            "<div class='notification error'>Version not found</div>",
            status_code=404,
        )

    # Get user from database
    user_result = await session.execute(select(User).where(User.keycloak_id == user.sub))
    db_user = user_result.scalar_one_or_none()
    user_id = db_user.id if db_user else None

    # Restore the data (this will create a new version)
    from sqlalchemy import func
    from sqlalchemy.orm.attributes import flag_modified

    # Get next version number
    max_result = await session.execute(
        select(func.coalesce(func.max(DatasetVersion.version_number), 0)).where(
            DatasetVersion.dataset_id == dataset_id
        )
    )
    max_version = max_result.scalar() or 0

    # Create new version with restored data
    new_version = DatasetVersion(
        dataset_id=dataset_id,
        version_number=max_version + 1,
        data=version.data,
        created_by_id=user_id,
    )
    session.add(new_version)

    # Update dataset
    dataset.data = version.data
    flag_modified(dataset, "data")
    session.add(dataset)
    await session.commit()

    # Clear cached state
    if dataset_id in dataset_states:
        del dataset_states[dataset_id]

    # Return redirect to reload page
    response = HTMLResponse(status_code=200)
    response.headers["HX-Redirect"] = f"/hub/datasets/{dataset_id}"
    return response
