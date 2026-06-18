"""Dataset create, import, delete, and example-loading routes."""

import copy
import logging
from pathlib import Path
from typing import Annotated, Any

from fastapi import File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from sqlalchemy import delete, select

from metaseed_hub.models import (
    ChatMessage,
    Comment,
    CommentReaction,
    Dataset,
    DatasetMember,
    DatasetVersion,
    Note,
    Spec,
    SpecDraft,
    SpecDraftMember,
    SpecStatus,
)
from metaseed_hub.ui.dependencies import (
    CurrentUser,
    DbSession,
    ensure_tenant_and_user,
    get_dataset_for_user,
)
from metaseed_hub.ui.helpers import (
    add_entity_node,
    create_nested_nodes,
    get_dataset_state,
    save_dataset_state,
)
from metaseed_hub.ui.render import render_template
from metaseed_hub.ui.security import csrf_error_response, validate_csrf_or_error

from ._router import router

logger = logging.getLogger("metaseed_hub")


@router.get("/new", response_class=HTMLResponse)
async def dataset_new(
    request: Request,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Return dataset creation form."""
    from metaseed.specs.loader import SpecLoader

    # Get or create tenant and user
    tenant, db_user = await ensure_tenant_and_user(session, user)

    # Get available profiles and versions from metaseed
    from pathlib import Path

    import metaseed

    loader = SpecLoader()
    examples_dir = Path(metaseed.__file__).parent / "examples"
    profiles_data = []
    for profile_name in loader.list_profiles():
        versions = loader.list_versions(profile_name)

        # Sort versions in descending order (newest first)
        def version_key(v: str) -> tuple[int, ...]:
            try:
                return tuple(int(x) for x in v.split("."))
            except ValueError:
                return (0,)

        versions = sorted(versions, key=version_key, reverse=True)
        # Get profile metadata from latest version
        display_name = profile_name
        description = ""
        root_entity = "Investigation"
        if versions:
            try:
                spec = loader.load_profile(versions[0], profile_name)
                display_name = spec.display_name or profile_name
                description = spec.description or ""
                root_entity = spec.root_entity or "Investigation"
            except Exception:
                pass

        # Check if examples exist for this profile (check latest version)
        has_example = False
        if versions:
            example_path = examples_dir / profile_name / versions[0]
            has_example = example_path.exists() and any(example_path.glob("*.yaml"))

        profiles_data.append(
            {
                "name": profile_name,
                "display_name": display_name,
                "description": description,
                "root_entity": root_entity,
                "versions": versions,
                "latest_version": versions[0] if versions else "",
                "source": "builtin",
                "has_example": has_example,
            }
        )

    # Get spec drafts from this tenant
    drafts_result = await session.execute(select(SpecDraft).where(SpecDraft.tenant_id == tenant.id))
    owned_drafts = list(drafts_result.scalars().all())

    # Also get specs shared with this user via SpecDraftMember
    shared_drafts: list[SpecDraft] = []
    shared_result = await session.execute(
        select(SpecDraft)
        .join(SpecDraftMember, SpecDraftMember.spec_draft_id == SpecDraft.id)
        .where(SpecDraftMember.user_id == db_user.id)
    )
    shared_drafts = list(shared_result.scalars().all())

    # Combine and deduplicate
    seen_ids: set[str] = set()
    drafts: list[SpecDraft] = []
    for draft in owned_drafts + shared_drafts:
        if draft.id not in seen_ids:
            seen_ids.add(draft.id)
            drafts.append(draft)

    for draft in drafts:
        if draft.name:
            spec_data = draft.spec_data or {}
            profiles_data.append(
                {
                    "name": f"draft:{draft.id}",
                    "display_name": f"{draft.name} (Draft)",
                    "description": spec_data.get("description", ""),
                    "root_entity": spec_data.get("root_entity", "Investigation"),
                    "versions": [draft.version],
                    "latest_version": draft.version,
                    "source": "draft",
                }
            )

    # Published specs owned by this tenant, offered as starting points.
    specs_result = await session.execute(
        select(Spec)
        .where(
            Spec.tenant_id == tenant.id,
            Spec.deleted_at.is_(None),
            Spec.status == SpecStatus.PUBLISHED,
        )
        .order_by(Spec.updated_at.desc())
    )
    user_specs = list(specs_result.scalars().all())

    return render_template(
        request=request,
        name="dataset_new.html",
        context={
            "user": user,
            "tenant_id": tenant.id,
            "profiles": profiles_data,
            "user_specs": user_specs,
            "nav_active": "home",
        },
    )


@router.post("/import")
async def dataset_import(
    request: Request,
    session: DbSession,
    user: CurrentUser,
    file: Annotated[UploadFile, File()],
    name: Annotated[str, Form()],
    profile: Annotated[str | None, Form()] = None,
    version: Annotated[str | None, Form()] = None,
    csrf_token: Annotated[str | None, Form(alias="_csrf_token")] = None,
) -> RedirectResponse:
    """Import a dataset from an uploaded file (JSON, YAML, or Excel)."""
    import json

    import yaml
    from metaseed.specs.loader import SpecLoader

    from metaseed_hub.ui.helpers import validate_csrf_token

    if not validate_csrf_token(request, csrf_token):
        return RedirectResponse("/hub/?error=csrf_validation_failed", status_code=302)

    # Get or create tenant and user
    tenant, db_user = await ensure_tenant_and_user(session, user)

    # Read file content
    content = await file.read()
    filename = file.filename or ""

    # Parse based on file type. Keep the form-supplied profile/version; the
    # detection below only fills them in when the user did not select them.
    data = None

    try:
        if filename.endswith((".yaml", ".yml")):
            data = yaml.safe_load(content.decode("utf-8"))
        elif filename.endswith(".json"):
            data = json.loads(content.decode("utf-8"))
        elif filename.endswith((".xlsx", ".xls")):
            # Excel import - read sheets as entity data
            # Each sheet name = entity type, headers = field names
            from io import BytesIO

            import openpyxl

            wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)

            # Parse all sheets into entities dict
            entities_by_type: dict[str, list[dict[str, Any]]] = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    continue

                headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

                # Parse data rows (skip header and placeholder row if it has <field> format)
                for row in rows[1:]:
                    # Skip placeholder rows
                    first_val = str(row[0]) if row[0] else ""
                    if first_val.startswith("<") and first_val.endswith(">"):
                        continue

                    entity_data = {}
                    for i, val in enumerate(row):
                        if i < len(headers) and val is not None:
                            # Skip placeholder values
                            str_val = str(val)
                            if str_val.startswith("<") and str_val.endswith(">"):
                                continue
                            entity_data[headers[i]] = val

                    if entity_data:
                        if sheet_name not in entities_by_type:
                            entities_by_type[sheet_name] = []
                        entities_by_type[sheet_name].append(entity_data)

            # Store entities for later processing
            data = {"_entities_by_type": entities_by_type}

            # Use first entity of first sheet for root if available
            if entities_by_type:
                first_type = list(entities_by_type.keys())[0]
                if entities_by_type[first_type]:
                    data.update(entities_by_type[first_type][0])
        else:
            return RedirectResponse("/hub/datasets/new?error=unsupported_format", status_code=302)

        if not data:
            return RedirectResponse("/hub/datasets/new?error=empty_file", status_code=302)

        # Use provided profile/version, or try to detect from data
        if not profile and isinstance(data, dict):
            profile = data.get("profile") or data.get("_profile")
        if not version and isinstance(data, dict):
            version = data.get("version") or data.get("_version")

        # Default to miappe if not detected
        if not profile:
            profile = "miappe"
        if not version:
            loader = SpecLoader()
            versions = loader.list_versions(profile)
            version = versions[0] if versions else "1.1"

    except Exception as e:
        logger.exception(f"Failed to parse import file: {e}")
        return RedirectResponse("/hub/datasets/new?error=parse_error", status_code=302)

    # Create dataset
    dataset = Dataset(
        tenant_id=tenant.id,
        name=name,
        profile=profile,
        version=version,
        data={},
    )
    session.add(dataset)
    await session.commit()
    await session.refresh(dataset)

    # Try to import entities from data
    try:
        loader = SpecLoader(profile=profile)
        spec = loader.load_profile(version, profile)
        root_entity = spec.root_entity or "Investigation"

        state = get_dataset_state(dataset)
        state.reset()
        state.profile = profile
        state.version = version
        state.facade = None
        facade = state.get_or_create_facade()

        # Handle different data structures
        entities_by_type = data.get("_entities_by_type", {}) if isinstance(data, dict) else {}

        if entities_by_type:
            # Excel import - create entities by type
            # Process root entity first, then children in hierarchy order
            entity_order = [root_entity] + [e for e in facade.entities if e != root_entity]

            for entity_type in entity_order:
                if entity_type not in entities_by_type:
                    continue

                for entity_data in entities_by_type[entity_type]:
                    try:
                        # Filter out empty values
                        clean_data = {
                            k: v for k, v in entity_data.items() if v is not None and str(v).strip()
                        }
                        if clean_data:
                            node = add_entity_node(state, entity_type, clean_data)
                            if not state.editing_node_id:
                                state.editing_node_id = node.id
                    except Exception as e:
                        logger.warning(f"Failed to create {entity_type}: {e}")

        else:
            # JSON/YAML import - use existing logic
            entities_data = data.get("entities", []) if isinstance(data, dict) else []
            if not entities_data and isinstance(data, dict):
                # Try to use the data directly as root entity
                # Filter out metadata fields
                entity_data = {
                    k: v
                    for k, v in data.items()
                    if not k.startswith("_") and k not in ("profile", "version")
                }
                if entity_data:
                    node = add_entity_node(state, root_entity, entity_data)
                    state.editing_node_id = node.id
                    create_nested_nodes(
                        state, facade, node, root_entity, copy.deepcopy(entity_data)
                    )

        # Save to database with version history
        await save_dataset_state(session, dataset, state)

    except Exception as e:
        logger.warning(f"Could not import entities, dataset created empty: {e}")

    return RedirectResponse(f"/hub/datasets/{dataset.id}", status_code=303)


@router.post("")
async def dataset_create(
    request: Request,
    session: DbSession,
    user: CurrentUser,
    name: Annotated[str, Form()],
    profile: Annotated[str, Form()],
    version: Annotated[str, Form()],
    csrf_token: Annotated[str | None, Form(alias="_csrf_token")] = None,
    load_example: Annotated[str | None, Form()] = None,
) -> RedirectResponse:
    """Create a new dataset."""
    import metaseed
    import yaml
    from metaseed.specs.loader import SpecLoader

    from metaseed_hub.ui.helpers import validate_csrf_token

    if not validate_csrf_token(request, csrf_token):
        return RedirectResponse("/hub/?error=csrf_validation_failed", status_code=302)

    # Get or create tenant and user
    tenant, db_user = await ensure_tenant_and_user(session, user)

    # Check if using a draft spec
    spec_draft_id = None
    if profile.startswith("draft:"):
        spec_draft_id = profile.replace("draft:", "")
        # Get the draft to use its name as the profile
        draft_result = await session.execute(select(SpecDraft).where(SpecDraft.id == spec_draft_id))
        draft = draft_result.scalar_one_or_none()
        if draft:
            profile = draft.name.lower()  # Lowercase to match ProfileFacade behavior
            version = draft.version

    dataset = Dataset(
        tenant_id=tenant.id,
        name=name,
        profile=profile,
        version=version,
        spec_draft_id=spec_draft_id,
        data={},
    )
    session.add(dataset)
    await session.commit()
    await session.refresh(dataset)

    # Load example data if requested
    logger.info(
        f"dataset_create: load_example={load_example!r}, profile={profile}, version={version}"
    )
    if load_example == "true":
        examples_dir = Path(metaseed.__file__).parent / "examples"
        version_dir = examples_dir / profile / version
        yaml_files = list(version_dir.glob("*.yaml")) if version_dir.exists() else []

        if yaml_files:
            try:
                example_data = yaml.safe_load(yaml_files[0].read_text(encoding="utf-8"))
                # Deep copy to prevent Pydantic from modifying the original dict
                example_data_copy = copy.deepcopy(example_data)

                loader = SpecLoader(profile=profile)
                spec = loader.load_profile(version, profile)
                root_entity = spec.root_entity or "Investigation"

                state = get_dataset_state(dataset)
                state.reset()
                state.profile = profile
                state.version = version
                state.facade = None
                facade = state.get_or_create_facade()

                node = add_entity_node(state, root_entity, example_data)
                state.editing_node_id = node.id

                # Create nested child nodes from the unmodified copy
                create_nested_nodes(state, facade, node, root_entity, example_data_copy)

                # Save to database with version history
                await save_dataset_state(session, dataset, state)
            except Exception as e:
                logger.exception(f"Failed to load example data: {e}")

    return RedirectResponse(f"/hub/datasets/{dataset.id}", status_code=303)


@router.delete("/{dataset_id}", response_class=HTMLResponse)
async def dataset_delete(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Delete a dataset."""
    try:
        validate_csrf_or_error(request)
    except Exception:
        return csrf_error_response()

    try:
        # Verify user has access to this dataset
        dataset = await get_dataset_for_user(dataset_id, session, user)
    except Exception as e:
        logger.error(f"Failed to get dataset {dataset_id}: {e}")
        response = Response(status_code=200)
        response.headers["HX-Trigger"] = (
            '{"showToast": {"message": "Dataset not found or access denied", "type": "error"}}'
        )
        return response

    try:
        # Manually delete related records (in case CASCADE not set in DB)
        await session.execute(
            delete(CommentReaction).where(
                CommentReaction.comment_id.in_(
                    select(Comment.id).where(Comment.dataset_id == dataset_id)
                )
            )
        )
        await session.execute(delete(Comment).where(Comment.dataset_id == dataset_id))
        await session.execute(delete(ChatMessage).where(ChatMessage.dataset_id == dataset_id))
        await session.execute(delete(Note).where(Note.dataset_id == dataset_id))
        await session.execute(delete(DatasetMember).where(DatasetMember.dataset_id == dataset_id))
        await session.execute(delete(DatasetVersion).where(DatasetVersion.dataset_id == dataset_id))

        await session.delete(dataset)
        await session.commit()
    except Exception as e:
        logger.error(f"Failed to delete dataset {dataset_id}: {e}", exc_info=True)
        await session.rollback()
        error_msg = str(e).replace('"', '\\"')
        response = Response(status_code=200)
        response.headers["HX-Trigger"] = (
            f'{{"showToast": {{"message": "Delete failed: {error_msg}", "type": "error"}}}}'
        )
        return response

    # Always redirect to home after delete
    response = Response(
        content='<script>window.location.href="/hub/";</script>',
        status_code=200,
        media_type="text/html",
    )
    response.headers["HX-Redirect"] = "/hub/"
    return response


@router.post("/{dataset_id}/load-example", response_class=HTMLResponse)
async def dataset_load_example(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Load example data into a dataset from YAML files."""
    import metaseed
    import yaml
    from metaseed.specs.loader import SpecLoader

    try:
        validate_csrf_or_error(request)
    except Exception:
        return csrf_error_response()

    # Verify user has access to this dataset
    dataset = await get_dataset_for_user(dataset_id, session, user)

    # Find example YAML file
    examples_dir = Path(metaseed.__file__).parent / "examples"
    version_dir = examples_dir / dataset.profile / dataset.version

    if not version_dir.exists():
        msg = f"No example available for {dataset.profile} v{dataset.version}"
        return HTMLResponse(f"<div class='error'>{msg}</div>")

    yaml_files = list(version_dir.glob("*.yaml"))
    if not yaml_files:
        msg = f"No example file found for {dataset.profile} v{dataset.version}"
        return HTMLResponse(f"<div class='error'>{msg}</div>")

    example_file = yaml_files[0]
    try:
        example_data = yaml.safe_load(example_file.read_text(encoding="utf-8"))
    except yaml.YAMLError as e:
        return HTMLResponse(f"<div class='error'>Error loading example: {e}</div>")

    # Deep copy to prevent Pydantic from modifying the original dict
    example_data_copy = copy.deepcopy(example_data)

    # Load spec to get root entity
    loader = SpecLoader(profile=dataset.profile)
    spec = loader.load_profile(dataset.version, dataset.profile)
    root_entity = spec.root_entity or "Investigation"

    # Load example into dataset state (append, don't replace)
    state = get_dataset_state(dataset)
    facade = state.get_or_create_facade()

    try:
        node = add_entity_node(state, root_entity, example_data)
        state.editing_node_id = node.id

        # Create nested child nodes from the unmodified copy
        create_nested_nodes(state, facade, node, root_entity, example_data_copy)

        # Save to database with version history
        logger.info(f"Saving tree with {len(state.entity_tree)} root nodes")
        for n in state.entity_tree:
            if n.instance:
                data = n.instance.model_dump(exclude_none=True)
                logger.info(f"  {n.entity_type} '{n.label}': {len(data)} fields")
            for c in n.children[:3]:
                if c.instance:
                    cdata = c.instance.model_dump(exclude_none=True)
                    logger.info(f"    Child {c.entity_type} '{c.label}': {len(cdata)} fields")

        await save_dataset_state(session, dataset, state)

    except Exception as e:
        import traceback

        logger.exception(f"Failed to load example: {e}")
        tb = traceback.format_exc()
        error_html = f"""
        <div class='notification error' style='user-select: text;'>
            <strong>Error loading example:</strong>
            <pre style='white-space: pre-wrap; font-size: 0.75rem; margin-top: 0.5rem;'>{e}

{tb}</pre>
        </div>"""
        return HTMLResponse(error_html)

    # Use HX-Redirect for HTMX to do a full page redirect
    response = HTMLResponse(status_code=200)
    response.headers["HX-Redirect"] = f"/hub/datasets/{dataset_id}"
    return response
