"""Dataset editor, tree, overview, validation, graph, and export routes."""

import logging
from typing import Annotated, Any

from fastapi import File, Request, UploadFile
from fastapi.responses import HTMLResponse, Response
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from metaseed_hub.models import (
    Dataset,
    DatasetMember,
    Tenant,
)
from metaseed_hub.ui.dependencies import (
    CurrentUser,
    DbSession,
    get_dataset_for_user,
)
from metaseed_hub.ui.helpers import (
    add_entity_node,
    ensure_dataset_facade,
    get_tree_data_from_nodes,
    save_dataset_state,
)
from metaseed_hub.ui.render import render_template
from metaseed_hub.ui.security import csrf_error_response, validate_csrf_or_error

from ._router import router

logger = logging.getLogger("metaseed_hub")


async def _build_dataset_context(
    dataset: Dataset,
    session: Any,
) -> dict[str, Any]:
    """Build common context for dataset views.

    Args:
        dataset: Dataset model.
        session: Database session for loading spec drafts.

    Returns:
        Dictionary with state, tree_data, and entity_descriptions.
    """
    state = await ensure_dataset_facade(dataset, session)
    tree_data = get_tree_data_from_nodes(state)

    # Get descriptions for entity types
    entity_descriptions: dict[str, str] = {}

    try:
        facade = state.get_or_create_facade()
        for entity_name in facade.entities:
            helper = getattr(facade, entity_name, None)
            if helper:
                entity_descriptions[entity_name] = helper.description or ""

    except Exception as e:
        # Log but don't crash
        logger.warning(f"Failed to load facade for dataset {dataset.id}: {e}")

    return {
        "state": state,
        "tree_data": tree_data,
        "entity_descriptions": entity_descriptions,
    }


@router.get("/{dataset_id}", response_class=HTMLResponse)
async def dataset_editor(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Dataset editor - wraps metaseed UI with hub chrome."""
    # Verify user has access to this dataset
    dataset = await get_dataset_for_user(dataset_id, session, user)

    # Load tenant for breadcrumb
    tenant = await session.get(Tenant, dataset.tenant_id)

    # Load members with user info
    members_result = await session.execute(
        select(DatasetMember)
        .where(DatasetMember.dataset_id == dataset_id)
        .options(selectinload(DatasetMember.user))
    )
    members = list(members_result.scalars().all())

    # Build common dataset context
    try:
        ctx = await _build_dataset_context(dataset, session)
        root_types = ctx["state"].get_root_entity_types()
    except Exception as e:
        # Profile doesn't exist or is invalid - show error page
        logger.warning(f"Failed to load dataset {dataset_id}: {e}")
        return render_template(
            request=request,
            name="dataset.html",
            context={
                "user": user,
                "dataset": dataset,
                "tenant": tenant,
                "members": members,
                "state": None,
                "root_types": [],
                "tree_data": [],
                "entity_descriptions": {},
                "nav_active": "home",
                "error": (
                    f"Could not load profile '{dataset.profile}' v{dataset.version}. "
                    "The profile may not exist or may be invalid."
                ),
            },
        )

    return render_template(
        request=request,
        name="dataset.html",
        context={
            "user": user,
            "dataset": dataset,
            "tenant": tenant,
            "members": members,
            "state": ctx["state"],
            "root_types": root_types,
            "tree_data": ctx["tree_data"],
            "entity_descriptions": ctx["entity_descriptions"],
            "nav_active": "home",
        },
    )


@router.get("/{dataset_id}/tree", response_class=HTMLResponse)
async def dataset_tree(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Return entity tree for a dataset."""
    # Verify user has access to this dataset
    dataset = await get_dataset_for_user(dataset_id, session, user)

    state = await ensure_dataset_facade(dataset, session)
    tree_data = get_tree_data_from_nodes(state)

    return render_template(
        request=request,
        name="partials/entity_tree.html",
        context={
            "tree_data": tree_data,
            "dataset_id": dataset_id,
        },
    )


@router.get("/{dataset_id}/root-buttons", response_class=HTMLResponse)
async def dataset_root_buttons(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Return root entity type buttons for the sidebar."""
    dataset = await get_dataset_for_user(dataset_id, session, user)

    state = await ensure_dataset_facade(dataset, session)
    root_types = state.get_root_entity_types()
    tree_data = get_tree_data_from_nodes(state)
    existing_root_types = [item["entity_type"] for item in tree_data]

    return render_template(
        request=request,
        name="partials/root_buttons.html",
        context={
            "dataset_id": dataset_id,
            "root_types": root_types,
            "existing_root_types": existing_root_types,
        },
    )


@router.get("/{dataset_id}/overview", response_class=HTMLResponse)
async def dataset_overview(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Return the overview panel with version history, comments, and sharing."""
    dataset = await get_dataset_for_user(dataset_id, session, user)

    # Load members
    members_result = await session.execute(
        select(DatasetMember)
        .where(DatasetMember.dataset_id == dataset_id)
        .options(selectinload(DatasetMember.user))
    )
    members = list(members_result.scalars().all())

    # Check if there's any tree data
    state = await ensure_dataset_facade(dataset, session)
    tree_data = get_tree_data_from_nodes(state)

    return render_template(
        request=request,
        name="partials/dataset_overview.html",
        context={
            "dataset": dataset,
            "members": members,
            "tree_data": tree_data,
        },
    )


@router.post("/{dataset_id}/validate", response_class=HTMLResponse)
async def dataset_validate(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
) -> HTMLResponse:
    """Validate all entities in the dataset against their schemas."""
    import html as html_module

    from pydantic import ValidationError

    try:
        validate_csrf_or_error(request)
    except Exception:
        return csrf_error_response()

    # Verify user has access to this dataset
    dataset = await get_dataset_for_user(dataset_id, session, user)

    state = await ensure_dataset_facade(dataset, session)
    facade = state.get_or_create_facade()

    errors: list[dict[str, Any]] = []

    for node_id, node in state.nodes_by_id.items():
        try:
            helper = getattr(facade, node.entity_type)
            data = node.instance.model_dump(exclude_none=True) if node.instance else {}
            # Re-validate by recreating - Pydantic validation runs here
            helper.create(**data)
        except ValidationError as e:
            errors.append(
                {
                    "node_id": node_id,
                    "entity_type": node.entity_type,
                    "label": node.label,
                    "errors": [
                        {"field": ".".join(str(x) for x in err["loc"]), "message": err["msg"]}
                        for err in e.errors()
                    ],
                }
            )
        except AttributeError:
            errors.append(
                {
                    "node_id": node_id,
                    "entity_type": node.entity_type,
                    "label": node.label,
                    "errors": [
                        {"field": "", "message": f"Unknown entity type: {node.entity_type}"}
                    ],
                }
            )
        except Exception as e:
            errors.append(
                {
                    "node_id": node_id,
                    "entity_type": node.entity_type,
                    "label": node.label,
                    "errors": [{"field": "", "message": str(e)}],
                }
            )

    # Count entities by type
    entity_counts: dict[str, int] = {}
    for node in state.nodes_by_id.values():
        entity_counts[node.entity_type] = entity_counts.get(node.entity_type, 0) + 1

    # Build HTML response
    total = len(state.nodes_by_id)
    valid_count = total - len(errors)

    html = '<div class="validation-results">'

    # Summary section
    if not errors:
        html += f"""
        <div class="validation-summary validation-success">
            <div class="validation-icon">&#10003;</div>
            <div class="validation-summary-text">
                <strong>All {total} entities are valid</strong>
                <p>Your dataset passes all validation checks.</p>
            </div>
        </div>
        """
    else:
        html += f"""
        <div class="validation-summary validation-error">
            <div class="validation-icon">&#10007;</div>
            <div class="validation-summary-text">
                <strong>{len(errors)} of {total} entities have issues</strong>
                <p>{valid_count} valid. Fix issues below.</p>
            </div>
        </div>
        """

    # Entity type breakdown
    html += '<div class="validation-breakdown"><h4>Entity Summary</h4><div class="breakdown-grid">'
    for entity_type, count in sorted(entity_counts.items()):
        error_count = sum(1 for e in errors if e["entity_type"] == entity_type)
        status_class = "valid" if error_count == 0 else "invalid"
        html += f"""
        <div class="breakdown-item {status_class}">
            <span class="breakdown-type">{html_module.escape(entity_type)}</span>
            <span class="breakdown-count">{count - error_count}/{count} valid</span>
        </div>
        """
    html += "</div></div>"

    # Detailed errors
    if errors:
        html += '<div class="validation-errors"><h4>Issues to Fix</h4>'
        for err in errors:
            html += f"""
            <div class="validation-error-item">
                <div class="validation-entity">
                    <span class="entity-type-badge">{html_module.escape(err["entity_type"])}</span>
                    <a href="#" class="entity-link"
                       hx-get="/hub/datasets/{dataset_id}/entity/{err["node_id"]}"
                       hx-target="#editor"
                       hx-swap="innerHTML">{html_module.escape(err["label"] or "")}</a>
                </div>
                <ul class="validation-error-list">
            """
            for field_err in err["errors"]:
                field = field_err["field"] or "(general)"
                html += (
                    f"<li><code>{html_module.escape(field)}</code>: "
                    f"{html_module.escape(field_err['message'])}</li>"
                )
            html += "</ul></div>"
        html += "</div>"

    html += "</div>"

    return HTMLResponse(html)


@router.get("/{dataset_id}/graph", response_class=HTMLResponse)
async def dataset_graph(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
    node_id: str | None = None,
) -> Response:
    """Graph visualization of dataset entities.

    Args:
        node_id: Optional. If provided, only show this entity and its descendants.
    """
    # Verify user has access to this dataset
    dataset = await get_dataset_for_user(dataset_id, session, user)

    return render_template(
        request=request,
        name="graph.html",
        context={
            "user": user,
            "dataset": dataset,
            "node_id": node_id,
            "nav_active": "home",
        },
    )


@router.get("/{dataset_id}/api/graph")
async def dataset_graph_api(
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
    node_id: str | None = None,
) -> Response:
    """Return graph data for visualization (JSON API).

    Uses metaseed's build_graph to extract nodes and edges from instance data.

    Args:
        node_id: Optional. If provided, only include this node and its descendants.
    """
    from fastapi.responses import JSONResponse
    from metaseed.ui.services.graph import build_graph

    # Verify user has access to this dataset
    dataset = await get_dataset_for_user(dataset_id, session, user)

    state = await ensure_dataset_facade(dataset, session)

    # Use metaseed's graph builder which properly extracts nested entities
    graph_data = build_graph(state)

    return JSONResponse(content=graph_data)


@router.get("/{dataset_id}/export")
async def dataset_export(
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
) -> Response:
    """Export dataset data to Excel file.

    Uses metaseed's export service to generate an Excel workbook
    containing all entities in the dataset.
    """
    from fastapi.responses import StreamingResponse
    from metaseed.ui.services.export import export_to_bytes, generate_filename

    dataset = await get_dataset_for_user(dataset_id, session, user)
    state = await ensure_dataset_facade(dataset, session)

    # Generate Excel file using metaseed's export service
    excel_bytes = export_to_bytes(state)
    filename = generate_filename(state)

    # If no data, use dataset name for filename
    if not filename or filename == "export.xlsx":
        filename = f"{dataset.name.replace(' ', '_')}.xlsx"

    return StreamingResponse(
        excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/{dataset_id}/import")
async def dataset_import_into_existing(
    request: Request,
    dataset_id: str,
    session: DbSession,
    user: CurrentUser,
    file: Annotated[UploadFile, File()],
) -> Response:
    """Import data from file into an existing dataset.

    Supports JSON, YAML, and Excel files. Adds entities to the existing dataset.
    """
    import json
    from io import BytesIO

    import yaml

    try:
        validate_csrf_or_error(request)
    except Exception:
        return csrf_error_response()

    dataset = await get_dataset_for_user(dataset_id, session, user)

    # Read file content
    content = await file.read()
    filename = file.filename or ""

    # Parse based on file type
    entities_by_type: dict[str, list[dict[str, Any]]] = {}

    try:
        if filename.endswith((".yaml", ".yml")):
            data = yaml.safe_load(content.decode("utf-8"))
            # For YAML, treat as single root entity or check for entities list
            if isinstance(data, dict) and "entities" in data:
                for entity in data["entities"]:
                    etype = entity.get("_type", dataset.profile)
                    if etype not in entities_by_type:
                        entities_by_type[etype] = []
                    entities_by_type[etype].append(entity)
            elif isinstance(data, dict):
                # Single entity - use root entity type
                from metaseed.specs.loader import SpecLoader

                loader = SpecLoader(profile=dataset.profile)
                spec = loader.load_profile(dataset.version, dataset.profile)
                root_entity = spec.root_entity or "Investigation"
                entities_by_type[root_entity] = [data]

        elif filename.endswith(".json"):
            data = json.loads(content.decode("utf-8"))
            if isinstance(data, dict) and "entities" in data:
                for entity in data["entities"]:
                    etype = entity.get("_type", dataset.profile)
                    if etype not in entities_by_type:
                        entities_by_type[etype] = []
                    entities_by_type[etype].append(entity)
            elif isinstance(data, dict):
                from metaseed.specs.loader import SpecLoader

                loader = SpecLoader(profile=dataset.profile)
                spec = loader.load_profile(dataset.version, dataset.profile)
                root_entity = spec.root_entity or "Investigation"
                entities_by_type[root_entity] = [data]

        elif filename.endswith((".xlsx", ".xls")):
            import openpyxl

            wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    continue

                headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

                for row in rows[1:]:
                    first_val = str(row[0]) if row[0] else ""
                    if first_val.startswith("<") and first_val.endswith(">"):
                        continue

                    entity_data: dict[str, Any] = {}
                    for i, val in enumerate(row):
                        if i < len(headers) and val is not None:
                            str_val = str(val)
                            if str_val.startswith("<") and str_val.endswith(">"):
                                continue
                            entity_data[headers[i]] = val

                    if entity_data:
                        if sheet_name not in entities_by_type:
                            entities_by_type[sheet_name] = []
                        entities_by_type[sheet_name].append(entity_data)
        else:
            return HTMLResponse(
                "<div class='notification error'>Unsupported file format</div>",
                status_code=400,
            )

    except Exception as e:
        logger.exception(f"Failed to parse import file: {e}")
        return HTMLResponse(
            f"<div class='notification error'>Parse error: {e}</div>",
            status_code=400,
        )

    # Get dataset state and add entities
    state = await ensure_dataset_facade(dataset, session)
    facade = state.get_or_create_facade()

    imported_count = 0
    errors: list[str] = []

    # Process entities in order (root first)
    from metaseed.specs.loader import SpecLoader

    loader = SpecLoader(profile=dataset.profile)
    spec = loader.load_profile(dataset.version, dataset.profile)
    root_entity = spec.root_entity or "Investigation"

    entity_order = [root_entity] + [e for e in facade.entities if e != root_entity]

    for entity_type in entity_order:
        if entity_type not in entities_by_type:
            continue

        for entity_data in entities_by_type[entity_type]:
            try:
                clean_data = {
                    k: v
                    for k, v in entity_data.items()
                    if v is not None and str(v).strip() and not k.startswith("_")
                }
                if clean_data:
                    add_entity_node(state, entity_type, clean_data)
                    imported_count += 1
            except Exception as e:
                errors.append(f"{entity_type}: {e}")

    # Save to database with version history
    await save_dataset_state(session, dataset, state)

    if errors:
        msg = f"Imported {imported_count} entities with {len(errors)} errors"
        logger.warning(f"Import errors: {errors[:5]}")
    else:
        msg = f"Successfully imported {imported_count} entities"

    return HTMLResponse(f"<div class='notification success'>{msg}</div>")


# =============================================================================
# Member Management Routes
# =============================================================================
